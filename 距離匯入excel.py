import requests
import time
from openpyxl import Workbook,load_workbook
wb = load_workbook('109.xlsx')
ws = wb.active
print(ws['A2'].value)
while 7==7:
    response = requests.get('https://norgay.center/php/vilsnodes.php?loadranging=1&project_id=1')
    str1 = response.text.split('Tag1196')
    str2= str1[1].split('"')
    str3= str2[5].split('\\')
    x=int(str3[0])
    print(x)
    time.sleep(1.5)